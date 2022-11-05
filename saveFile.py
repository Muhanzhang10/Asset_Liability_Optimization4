from tree import tree
from openpyxl import load_workbook
from openpyxl import Workbook
#write the result to excel

def printVarhelper(self, index):
        name = ""
        temp = self
        while True:
            name += " " + temp.name 
            for i in range(len(temp.children)):
                if temp.children[i].range[1] > index:
                   temp = temp.children[i]
                   if temp.is_leaf() == True:
                       name += ": "
                       return name[1::]
                   break
        return "Failed finding corresponding name"

#print all the leaf nodes' path and their corresponding value
#write result in excel file
def printVariables(self, var, path, month = 1):
    wb = load_workbook(path)
    ws = wb.create_sheet(str(month)) # insert at the end (default)
    ws.title = str(month)
    ws.sheet_properties.tabColor = "1072BA"
    cell = ws.cell(row=4, column=2, value=10)
    cell.value = "Insert"
    for i in range(len(var)):
        name = printVarhelper(self, i)
        #print(name, var[i])
        list_name = name[:-2:].split()
        for col in range(len(list_name)):
            mycell = ws.cell(row = i + 1, column = col + 1)  
            mycell.value = list_name[col]
        mycell = ws.cell(row= i + 1, column = len(list_name) + 1)  
        mycell.value = var[i]

    wb.save(path)


def save(Tree, info, var, month = 1, path = '/Users/zhangmuhan/Desktop/result.xlsx'):
    printVariables(Tree, var, path, month)
    wb = load_workbook(path)
    ws = wb.create_sheet("NII" + str(month))
    ws.title = "NII" + str(month)
    ws.sheet_properties.tabColor = "1072BA"
    ws.cell(row = 1, column = 1).value = "Month: "
    ws.cell(row = 2, column = 1).value = 'Asset total: '
    ws.cell(row = 3, column = 1).value = 'Liability total: '
    ws.cell(row = 4, column = 1).value = "Before optimization: "
    ws.cell(row = 5, column = 1).value = "After optimization: "
    ws.cell(row = 1, column = 2).value = str(month)
    ws.cell(row = 2, column = 2).value = info['Asset total']
    ws.cell(row = 3, column = 2).value = info['Liability total']
    ws.cell(row = 4, column = 2).value = info['Before optimization']
    ws.cell(row = 5, column = 2).value = info['After optimization']
    wb.save(path)
