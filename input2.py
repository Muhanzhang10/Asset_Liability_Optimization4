from openpyxl import load_workbook
from tree import tree


def analyze(string):
    string = string.split('-')
    string0 = string[0]
    numl = [string0[:3:]]
    for i in range(3, len(string0)):
        if i%2 == 1:
            numl.append(string0[i: i + 2])
    path = numl.copy()
    path.insert(0, 'Root')
    if len(string) == 2:
        path.append(string[-1])
    return path
    

wb = load_workbook(r'/Users/zhangmuhan/Desktop/基础参数数据-20210623.xlsx')
ws = wb['Account维度']
Tree = tree('Root')
Tree.name2 = '资产负债产品'
for row in range(4, 100000):
    raw = ws.cell(row = row, column = 3).value
    if row == 4:
        Tree.add(ws.cell(row = row, column = 3).value, ['Root'])
        continue
    if raw == None:
        break
    path = analyze(raw)
    name = path[-1]
    Tree.add(name, path[:-1:])

Tree.add_range()
Tree.printTree()