from openpyxl import load_workbook
from tree import tree


layers = [['all'], ['asset', 'liability'], ['balance', 'interest'], ['month1', 'month2', 'month3', 'month4', 'month5', 'month6', 'month7', 'month8', 'month9', 'month10', 'month11', 'month12']]

Tree = tree('all')
for layer in layers:
    if layer == ['all']:
        continue
    Tree.add_multiples(layer)

Tree.children[0].add_multiples(['a1', 'a2', 'a3'])
Tree.children[0].add_multiples(['term1', 'term2', 'term3', 'term4', 'term5'])
#Tree.children[1].add_multiples(['l1', 'l2', 'l3', 'l4'])

l1 = tree('l1')
l2 = tree('l2')
l3 = tree('l3')
l4 = tree('l4')

l1.add_multiples(['term1', 'term2', 'term3', 'term4', 'term5', 'term6'])
l2.add_multiples(['term1', 'term2', 'term3', 'term4', 'term5', 'term6'])
l3.add_multiples(['term1'])
l4.add_multiples(['term1'])
trees = [l1, l2, l3, l4]
Tree.children[1].connect(trees)
Tree.add_range()




wb = load_workbook('场景2 model base.xlsx')
ws = wb['12个切片 balance']
var_balance_asset = []
var_balance_liability =[]
for month in range(1, 13):
    column = 3 + 2 * (month - 1)
    for row in range(2, 17):
        var_balance_asset.append(ws.cell(row = row, column = column).value)
    for row2 in range(17, 31):
        var_balance_liability.append(ws.cell(row = row2, column = column).value)
ws2 = wb['12个切片 interest']
var_interest_asset = []
var_interest_liability = []
for month in range(1, 13):
    column = 4 + 2 * (month - 1)
    for row in range(2, 17):
        var_interest_asset.append(ws2.cell(row = row, column = column).value)
    for row2 in range(17, 31):
        var_interest_liability.append(ws2.cell(row = row2, column = column).value)
var = var_balance_asset + var_interest_asset + var_balance_liability + var_interest_liability
#domain = Tree.getTo(['all', 'liability', 'balance', 'month12']).range
#print(var[domain[0]: domain[1]: ])

bound = []
for i in range(696):
    bound.append(0.05)


a = Tree.getTo(['all', 'asset', 'interest', 'month1']).range
l = Tree.getTo(['all', 'liability', 'interest', 'month1']).range
print(var[a[0]:a[1]:])
print(var[l[0]:l[1]:])